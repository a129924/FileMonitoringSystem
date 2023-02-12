import os
from openpyxl import load_workbook, Workbook

from regex_converter import CheckCommaList
from typing import List, Any

class ExcelDataParser():
    def __init__(self, file_path: str, sheet_name: str, is_convert = True) -> None:
        assert os.path.splitext(file_path)[1] in [".xls", ".xlsx", ".csv"], "檔案副檔名錯誤" 
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.is_convert = is_convert

    @property
    def wb(self)->Workbook:
        return load_workbook(self.file_path)
    
    @property
    def sheet_names(self,)->list:
        return self.wb.sheetnames

    @property
    def data(self)->list:
        data = []
        if self.is_convert:
            for rows in self.wb[self.sheet_name]:
                data.append(CheckCommaList([row.value for row in rows]))
        else:
            for rows in self.wb[self.sheet_name]:
                data.append([row.value for row in rows])
        return data

    @property
    def header(self)->list:
        return self.data[0]
    
    @property
    def body(self)->list:
        return self.data[1:]

    @property
    def dataframe(self,)->list:
        return [dict(zip(self.data[0], row)) for row in self.data[1:]]

    @staticmethod
    def reset_index(dataframe:list, index:str)->dict:
        assert index in dataframe[0].keys(), "Key不存在"
        reset_index_data = {}
        for row in dataframe:
            new_index = row[index]
            del row[index]
            reset_index_data[new_index] = row

        return reset_index_data
    
    def fillna(self, index, method):
        
        if method == "first":
            data = self.data
        elif method == "last":
            data = self.data[::-1]
        else:
            raise ValueError("Invalid method")
        
        fillna_value = data[0][index]
        
        for row in data[1:]:
            if row[index] != None:
                fillna_value = row[index]
                continue
            
            row[index] = fillna_value

        return data
    
    def groupby(self, index:int, data:List[List[str]])->dict:
        groupby_data = {}
        for row in data:
            if row[index] not in groupby_data:
                groupby_data[row[index]] = row[:index] + row[index+1:]
            else:
                # groupby_data[row[index]] = groupby_data[row[index]] + row[:index] + row[index+1:]
                groupby_data[row[index]] += row[:index] + row[index+1:]

        return groupby_data
    
if "__main__" == __name__:
    data = ExcelDataParser(
        r"D:\CODE\NEW\FileMonitoringSystem-main\Setting\ProjectSettingInfo.xlsx", "ProjectSettingInfo")
    print(list(data.reset_index(data.dataframe, "目的路徑").keys()))


